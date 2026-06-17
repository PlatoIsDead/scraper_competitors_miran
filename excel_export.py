"""Excel export — appends a new monthly auto-pivot sheet to the comparison workbook.

CLI:
    python excel_export.py            # scrape all + export
    python excel_export.py --cleanup  # remove "Лист 65" (idempotent, run once)
    python excel_export.py --no-scrape  # export from existing history.csv

The workbook must be closed in Excel before writing.
"""

import os
import re
import argparse
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from pivot_utils import build_pivot, PROVIDERS

DATA_DIR = "data"
XLSX_PATH = os.path.join(DATA_DIR, "Сравнение стоимости DS.xlsx")
HISTORY_CSV = os.path.join(DATA_DIR, "history.csv")

RU_MONTHS = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

# Column A–M headers matching the Март_26_SilverGold template exactly
HEADERS = [
    "CPU", "RAM", "HDD", "Internet",
    "Миран тарифы",
    "Миран по\n калькулятору",
    "Селектел\nhttps://selectel.ru/",
    "Рег.ру\nreg.ru",
    "1stDedic\nhttps://1dedic.ru/",
    "Hostkey\nhttps://hostkey.ru/",
    "IT-Lite\nit-lite.ru",
    "NetRack\nnetrack.ru",
    "Timeweb\ntimeweb.com",
]

# scraper provider key → 1-based column index (A=1 … M=13)
# F (col 6) = "Миран по калькулятору" — left empty (manual entry)
PROVIDER_COL: dict[str, int] = {
    "miran":    5,   # E
    "selectel": 7,   # G
    "regcloud": 8,   # H
    "1dedic":   9,   # I
    "hostkey":  10,  # J
    "it-lite":  11,  # K
    "netrack":  12,  # L
    "timeweb":  13,  # M
}

# Columns included in per-row min/max coloring (exclude F = calculator)
COLOR_COLS = [5, 7, 8, 9, 10, 11, 12, 13]

RED_FILL   = PatternFill(fgColor="FFF4CCCC", fill_type="solid")
GREEN_FILL = PatternFill(fgColor="FFD9EAD3", fill_type="solid")

HEADER_FONT = Font(bold=True)

COL_WIDTHS = {
    1: 42.5,   # A CPU
    2: 13.0,   # B RAM
    3: 25.5,   # C HDD
    4: 13.0,   # D Internet
}
DEFAULT_COL_WIDTH = 13.0


# ── Internal helpers ──────────────────────────────────────────────────

def _auto_sheet_name(wb, today: date) -> str:
    """Build 'Июнь_26_auto', appending _2/_3 on collision."""
    base = f"{RU_MONTHS[today.month]}_{today.year % 100:02d}_auto"
    name = base
    i = 2
    while name in wb.sheetnames:
        name = f"{base}_{i}"
        i += 1
    return name


def _color_row(ws, row: int) -> None:
    """Apply static green/red fills to the min/max price cells in a data row."""
    vals: dict[int, float] = {}
    for col in COLOR_COLS:
        v = ws.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            vals[col] = float(v)

    if len(vals) < 2:
        return

    lo = min(vals.values())
    hi = max(vals.values())
    if lo == hi:
        return  # all equal — nothing meaningful to highlight

    for col, v in vals.items():
        if v == lo:
            ws.cell(row=row, column=col).fill = GREEN_FILL
        elif v == hi:
            ws.cell(row=row, column=col).fill = RED_FILL


# ── Public API ────────────────────────────────────────────────────────

def export_to_excel(
    df: pd.DataFrame,
    xlsx_path: str = XLSX_PATH,
    today: date | None = None,
) -> str:
    """Append a new monthly auto-pivot sheet to the existing workbook.

    Returns the created sheet name.
    Preserves all existing sheets and their formatting.
    The workbook must NOT be open in Excel when this is called.
    """
    today = today or date.today()
    pivot = build_pivot(df, PROVIDERS)

    wb = load_workbook(xlsx_path)
    sheet_name = _auto_sheet_name(wb, today)
    ws = wb.create_sheet(sheet_name)

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)

    # Column widths
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            COL_WIDTHS.get(col_idx, DEFAULT_COL_WIDTH)
        )

    # Data rows
    for r_offset, (_, pivot_row) in enumerate(pivot.iterrows(), start=2):
        ws.cell(row=r_offset, column=1, value=pivot_row.get("CPU", ""))
        ws.cell(row=r_offset, column=2, value=pivot_row.get("RAM", ""))
        ws.cell(row=r_offset, column=3, value=pivot_row.get("Диск", ""))
        ws.cell(row=r_offset, column=4, value="")   # Internet — manual
        ws.cell(row=r_offset, column=6, value="")   # Миран по калькулятору — manual

        for provider, col_idx in PROVIDER_COL.items():
            price = pivot_row.get(provider)
            if pd.notna(price):
                cell = ws.cell(row=r_offset, column=col_idx, value=float(price))
                cell.number_format = "General"

        _color_row(ws, r_offset)

    wb.save(xlsx_path)
    return sheet_name


def remove_legacy_sheet(
    xlsx_path: str = XLSX_PATH,
    name: str = "Лист 65",
) -> bool:
    """Remove a sheet if present and save. Idempotent. Returns True if removed."""
    wb = load_workbook(xlsx_path)
    if name not in wb.sheetnames:
        return False
    wb.remove(wb[name])
    wb.save(xlsx_path)
    return True


# ── CLI ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Экспорт цен дедиков в Excel")
    parser.add_argument(
        "--cleanup",
        action="store_true",
        help='Удалить лист "Лист 65" (идемпотентно) и выйти',
    )
    parser.add_argument(
        "--no-scrape",
        action="store_true",
        help="Не скрейпить — использовать последний снапшот из history.csv",
    )
    args = parser.parse_args()

    if args.cleanup:
        removed = remove_legacy_sheet()
        if removed:
            print('Лист "Лист 65" удалён.')
        else:
            print('Лист "Лист 65" не найден (уже удалён).')
        return

    if args.no_scrape:
        if not os.path.exists(HISTORY_CSV):
            print(f"Файл {HISTORY_CSV} не найден. Запустите без --no-scrape.")
            return
        df_all = pd.read_csv(HISTORY_CSV)
        if df_all.empty:
            print("history.csv пустой.")
            return
        df = df_all[df_all["scraped_at"] == df_all["scraped_at"].max()]
    else:
        from dedicated_scraper import scrape_all, save_history
        df = scrape_all()
        save_history(df)

    if df.empty:
        print("Нет данных для экспорта.")
        return

    sheet = export_to_excel(df)
    print(f"Записано в лист: «{sheet}»  →  {XLSX_PATH}")


if __name__ == "__main__":
    main()
