# Разовый конвертер: data/Parser.xlsx клиента → config/miran_configs.json
# и config/disk_classes.json. После генерации файлы ведутся руками;
# конвертер нужен только для пересоздания при обновлении Parser.xlsx.
#
# Использование:
#   python excel_to_configs.py [--xlsx data/Parser.xlsx] [--out config/miran_configs.json]
#                              [--classes-out config/disk_classes.json] [--dry-run]
#
# Лист «Данные»: Кол-во CPU | CPU | RAM | HDD | Миран по калькулятору.
# Лист «Сопоставление дисков»: секции по типу диска (SSD / NVME / SATA),
# каждая строка = группа эквивалентных размеров (960 ГБ ≈ 1 ТБ);
# колонки «ГБ» и «ТБ» из шапки листа задают единицу измерения.

import argparse
import json
import logging
import re
from datetime import date
from pathlib import Path

import openpyxl

from dedicated_scraper import normalize_disk_gb, normalize_disk_type

log = logging.getLogger("excel_to_configs")

DEFAULT_OUT = Path("config") / "miran_configs.json"
DEFAULT_CLASSES_OUT = Path("config") / "disk_classes.json"

DATA_SHEET = "Данные"
MAPPING_SHEET = "Сопоставление дисков"


def default_source() -> Path:
    """Свежайший из data/Parser.xlsx | data/Parser.ods (клиент шлёт то и то)."""
    candidates = [p for p in (Path("data") / "Parser.xlsx",
                              Path("data") / "Parser.ods") if p.exists()]
    if not candidates:
        raise ValueError("Не найден data/Parser.xlsx или data/Parser.ods")
    return max(candidates, key=lambda p: p.stat().st_mtime)

# multiplication signs seen in the workbook: ×, x, X, х (cyrillic), *
_MULT = r"[×xXх*]"

_DISK_SEG_RE = re.compile(
    rf"(\d+)\s*{_MULT}\s*(\d+(?:[.,]\d+)?)\s*(ГБ|ТБ|GB|TB)?\s*(NVME|NVMe|SSD|HDD)?",
    re.I,
)


def parse_ref_disks(text: str) -> list[dict]:
    """'2 * 1 ТБ NVME + 2 * 2 ТБ SSD' → список пулов. Сегменты без цифр
    (VROC, RAID) пропускаются. Нет единицы («2 х 10 HDD») → ТБ при ≤32."""
    pools = []
    if not text or not isinstance(text, str):
        return pools
    for segment in text.split("+"):
        segment = segment.strip()
        if not segment or not re.search(r"\d", segment):
            continue
        m = _DISK_SEG_RE.search(segment)
        if not m:
            log.warning("Не удалось разобрать сегмент дисков: «%s»", segment)
            continue
        count = int(m.group(1))
        size = float(m.group(2).replace(",", "."))
        unit = (m.group(3) or "").upper()
        if unit in ("ТБ", "TB"):
            size *= 1000
        elif not unit:
            if size <= 32:
                size *= 1000
                log.warning(
                    "Сегмент «%s» без единицы измерения — предполагаю ТБ", segment
                )
            else:
                log.warning(
                    "Сегмент «%s» без единицы измерения — предполагаю ГБ", segment
                )
        pools.append({
            # тип — по всему сегменту: «SSD NVME» должен дать NVMe
            "disk_type": normalize_disk_type(segment),
            "disk_count": count,
            "disk_size_gb": normalize_disk_gb(int(size)),
        })
    return pools


def resolve_cpu(model_raw: str, cpu_aliases: dict) -> dict | None:
    """Модель из колонки CPU → {cpu_model, cpu_model_display,
    cpu_cores_per_socket} по cpu_specs. Ядер в Parser.xlsx нет — модель
    обязана быть в словаре. cpu_model — канон для матчинга,
    cpu_model_display — как написал клиент (для отчёта и интерфейса)."""
    model = re.sub(r"\s+", " ", str(model_raw).strip())
    if not model:
        return None
    spec = cpu_aliases.get(model.lower())
    if not spec:
        log.warning("Модель CPU не найдена в cpu_specs.json: «%s» — пропуск", model)
        return None
    return {
        "cpu_model": spec["canonical"],
        "cpu_model_display": model,
        "cpu_cores_per_socket": spec["cores"],
    }


def _load_sheet_rows(path: Path, sheet: str) -> list[tuple]:
    """Лист книги → список кортежей значений (None вместо пустых).
    .xlsx — openpyxl; .ods (клиент работает в LibreOffice) — pandas+odfpy."""
    if path.suffix.lower() == ".ods":
        import pandas as pd
        try:
            df = pd.read_excel(path, engine="odf", sheet_name=sheet, header=None)
        except (ValueError, KeyError) as e:
            raise ValueError(f"Лист «{sheet}» не найден в {path}") from e
        return [
            tuple(None if pd.isna(v) else v for v in row)
            for row in df.itertuples(index=False, name=None)
        ]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Лист «{sheet}» не найден в {path}")
        return [tuple(row) for row in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()


def convert(xlsx_path: Path, cpu_aliases: dict, sheet: str = DATA_SHEET) -> dict:
    """Лист «Данные» → структура miran_configs.json (с ценой Миран)."""
    configs = []
    skipped = 0
    seen_keys: dict[tuple, str] = {}
    for row_idx, row in enumerate(_load_sheet_rows(xlsx_path, sheet), start=1):
            sockets_cell, cpu_cell, ram_cell, disk_cell, price_cell = (
                list(row) + [None] * 5
            )[:5]
            # содержательная строка: есть модель, RAM числом и диски;
            # шапка, пустые строки и строки-поколения отсеиваются здесь
            if cpu_cell is None or not isinstance(ram_cell, (int, float)) \
                    or not disk_cell:
                skipped += 1
                continue
            cpu = resolve_cpu(cpu_cell, cpu_aliases)
            if not cpu:
                log.warning("Строка %d: CPU «%s» не распознан — пропуск",
                            row_idx, cpu_cell)
                continue
            pools = parse_ref_disks(str(disk_cell))
            if not pools:
                log.warning("Строка %d: не удалось разобрать диски «%s» — пропуск",
                            row_idx, disk_cell)
                continue
            sockets = int(sockets_cell) if isinstance(sockets_cell, (int, float)) else 1
            price = float(price_cell) if isinstance(price_cell, (int, float)) else None
            if price is None:
                log.warning("Строка %d: нет цены Миран (%s, %s ГБ, %s)",
                            row_idx, cpu["cpu_model"], int(ram_cell), disk_cell)
            key = (
                cpu["cpu_model"], sockets, int(ram_cell),
                tuple(sorted((p["disk_type"], p["disk_count"], p["disk_size_gb"])
                             for p in pools)),
            )
            if key in seen_keys:
                log.warning("Строка %d: дубликат конфигурации (уже %s) — пропуск",
                            row_idx, seen_keys[key])
                continue
            config_id = f"MIR-{len(configs) + 1:03d}"
            seen_keys[key] = config_id
            configs.append({
                "config_id": config_id,
                **cpu,
                "cpu_sockets": sockets,
                "ram_gb": int(ram_cell),
                "disk_pools": pools,
                "miran_price": price,
                "source_row": row_idx,
            })

    log.info("Лист «%s»: конфигураций %d, пропущено строк %d",
             sheet, len(configs), skipped)
    return {
        "generated_from": f"{xlsx_path.name} / {sheet}",
        "generated_at": date.today().isoformat(),
        "configs": configs,
    }


def convert_disk_classes(xlsx_path: Path, sheet: str = MAPPING_SHEET) -> dict:
    """Лист «Сопоставление дисков» → структура disk_classes.json.

    Строка с текстом = заголовок секции (тип диска), числовая строка = группа
    эквивалентных размеров. Единица колонки (ГБ/ТБ) — из первой строки листа.
    """
    rows = _load_sheet_rows(xlsx_path, sheet)
    if not rows:
        raise ValueError(f"{xlsx_path}: лист «{sheet}» пуст")

    # шапка: какие колонки в ГБ, какие в ТБ
    tb_cols = {i for i, v in enumerate(rows[0])
               if isinstance(v, str) and v.strip().upper() in ("ТБ", "TB")}

    groups = []
    current_type: str | None = None
    for row in rows[1:]:
        texts = [v for v in row if isinstance(v, str) and v.strip()]
        if texts:
            current_type = normalize_disk_type(" ".join(texts))
            continue
        sizes = set()
        for i, v in enumerate(row):
            if not isinstance(v, (int, float)):
                continue
            sizes.add(round(v * 1000) if i in tb_cols else int(v))
        if not sizes:
            continue
        if current_type is None:
            log.warning("Группа размеров %s до заголовка типа — пропуск",
                        sorted(sizes))
            continue
        groups.append({
            "disk_type": current_type,
            "sizes_gb": sorted(sizes),
        })

    log.info("Лист «%s»: групп эквивалентности %d", sheet, len(groups))
    return {
        "generated_from": f"{xlsx_path.name} / {sheet}",
        "generated_at": date.today().isoformat(),
        "groups": groups,
    }


def load_cpu_aliases(specs_path: Path = Path("config") / "cpu_specs.json") -> dict:
    cpu_aliases: dict = {}
    if specs_path.exists():
        raw = json.loads(specs_path.read_text(encoding="utf-8"))
        for key, item in raw.items():
            cpu_aliases[key.lower()] = item
            for alias in item.get("aliases", []):
                cpu_aliases.setdefault(alias.lower(), item)
    else:
        log.warning("%s не найден — канонизация моделей CPU отключена", specs_path)
    return cpu_aliases


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Конвертация Parser.xlsx в config/miran_configs.json "
                    "и config/disk_classes.json"
    )
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="источник (xlsx/ods); по умолчанию — свежайший "
                             "из data/Parser.xlsx и data/Parser.ods")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--classes-out", type=Path, default=DEFAULT_CLASSES_OUT)
    parser.add_argument("--dry-run", action="store_true",
                        help="показать результат без записи файлов")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")

    source = args.xlsx or default_source()
    log.info("Источник эталона: %s", source)
    result = convert(source, load_cpu_aliases())
    classes = convert_disk_classes(source)

    if args.dry_run:
        for cfg in result["configs"]:
            pools_txt = " + ".join(
                f"{p['disk_count']}×{p['disk_size_gb']} ГБ {p['disk_type']}"
                for p in cfg["disk_pools"]
            )
            price_txt = (f"{cfg['miran_price']:.0f} ₽"
                         if cfg["miran_price"] is not None else "—")
            print(f"{cfg['config_id']} (строка {cfg['source_row']}): "
                  f"{cfg['cpu_sockets']} × {cfg['cpu_model']} "
                  f"({cfg['cpu_cores_per_socket']} ядер), "
                  f"{cfg['ram_gb']} ГБ, {pools_txt}, Миран {price_txt}")
        print(f"\nВсего: {len(result['configs'])} конфигураций, "
              f"{len(classes['groups'])} групп дисков (файлы не записаны)")
        return

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Записано {len(result['configs'])} конфигураций в {args.out}")
    args.classes_out.write_text(
        json.dumps(classes, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Записано {len(classes['groups'])} групп дисков в {args.classes_out}")


if __name__ == "__main__":
    main()
