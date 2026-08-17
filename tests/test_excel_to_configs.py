import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_to_configs import (
    convert,
    convert_disk_classes,
    parse_ref_disks,
    resolve_cpu,
)

ALIASES = {
    "intel xeon silver 4214r": {"canonical": "Intel Xeon Silver 4214R", "cores": 12},
    "silver 4314": {"canonical": "Intel Xeon Silver 4314", "cores": 16},
    "gold 5317": {"canonical": "Intel Xeon Gold 5317", "cores": 12},
    "e3-1231v3": {"canonical": "Intel Xeon E3-1231v3", "cores": 4},
    "intel xeon e-2386g": {"canonical": "Intel Xeon E-2386G", "cores": 6},
}


class TestResolveCpu:
    def test_full_model(self):
        cpu = resolve_cpu("Intel Xeon Silver 4214R", ALIASES)
        assert cpu == {
            "cpu_model": "Intel Xeon Silver 4214R",
            "cpu_cores_per_socket": 12,
        }

    def test_short_alias(self):
        assert resolve_cpu("Silver 4314", ALIASES)["cpu_model"] == \
            "Intel Xeon Silver 4314"

    def test_extra_whitespace_collapsed(self):
        assert resolve_cpu("  Gold   5317 ", ALIASES)["cpu_model"] == \
            "Intel Xeon Gold 5317"

    def test_unknown_model_returns_none(self):
        assert resolve_cpu("3 покеоление Intel", ALIASES) is None

    def test_empty_returns_none(self):
        assert resolve_cpu("", ALIASES) is None


class TestParseRefDisks:
    def test_simple_gb_ssd(self):
        assert parse_ref_disks("2 * 960 ГБ SSD") == [
            {"disk_type": "SSD", "disk_count": 2, "disk_size_gb": 1000}
        ]  # 960 близко к 1000 (≤5%) — снапится

    def test_decimal_tb_stays(self):
        # 1.9 ТБ = 1900: до 2000 больше 5% — размер сохраняется,
        # эквивалентность 1900 ≈ 2000 решает disk_classes.json
        assert parse_ref_disks("2 * 1.9 ТБ SSD") == [
            {"disk_type": "SSD", "disk_count": 2, "disk_size_gb": 1900}
        ]

    def test_sata_maps_to_hdd(self):
        assert parse_ref_disks("2 x 3 ТБ SATA") == [
            {"disk_type": "HDD", "disk_count": 2, "disk_size_gb": 3000}
        ]

    def test_no_space_tb_nvme(self):
        assert parse_ref_disks("2 * 1ТБ NVME") == [
            {"disk_type": "NVMe", "disk_count": 2, "disk_size_gb": 1000}
        ]

    def test_vroc_suffix_ignored(self):
        pools = parse_ref_disks("2 * 1 ТБ NVME + VROC")
        assert pools == [{"disk_type": "NVMe", "disk_count": 2, "disk_size_gb": 1000}]

    def test_multi_pool(self):
        pools = parse_ref_disks("2 * 1 ТБ NVME + 2 * 2 ТБ SSD")
        assert len(pools) == 2
        assert pools[0]["disk_type"] == "NVMe"
        assert pools[1] == {"disk_type": "SSD", "disk_count": 2, "disk_size_gb": 2000}

    def test_latin_tb(self):
        assert parse_ref_disks("2 * 2 TB NVME")[0]["disk_size_gb"] == 2000

    def test_missing_unit_10_means_tb(self):
        # кириллическая «х» и отсутствующая единица (10 → ТБ), без снапа
        pools = parse_ref_disks("2 х 10 HDD")
        assert pools == [{"disk_type": "HDD", "disk_count": 2, "disk_size_gb": 10000}]

    def test_empty(self):
        assert parse_ref_disks("") == []


@pytest.fixture
def workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Данные"
    rows = [
        ["Кол-во CPU", "CPU", "RAM", "HDD", "Миран по калькулятору"],
        [None, None, None, None, None],
        [1, "E3-1231v3", 16, "2 * 480 ГБ SSD", 8200],
        [2, "Intel Xeon Silver 4214R", 64, "2 * 960 ГБ SSD", 21000],
        [None, "3 покеоление Intel", None, None, None],
        [2, "Silver 4314", 128, "2 * 480 ГБ SSD + 2 * 2 ТБ SSD", 35000],
        [1, "Intel Xeon E-2386G", 64, "2 * 1 ТБ NVME", None],
        [2, "Intel Xeon Silver 4214R", 64, "2 * 960 ГБ SSD", 22222],
    ]
    for row in rows:
        ws.append(row)

    mapping = wb.create_sheet("Сопоставление дисков")
    map_rows = [
        ["ГБ", "ГБ", None, "ТБ", "ТБ"],
        ["SSD", "SSD SATA", None, None, None],
        [960, 1000, None, 0.96, 1],
        [None, None, None, None, None],
        ["NVME", "SSD NVME", None, None, None],
        [375, None, None, None, None],
        ["SATA", "HDD", None, None, None],
        [1000, None, None, 1, None],
    ]
    for row in map_rows:
        mapping.append(row)

    path = tmp_path / "Parser.xlsx"
    wb.save(path)
    return path


class TestConvert:
    def test_convert_produces_expected_configs(self, workbook):
        result = convert(workbook, ALIASES)
        configs = result["configs"]
        # шапка/пустая/поколение отсеяны, последняя строка — дубликат
        assert len(configs) == 4
        assert configs[0]["config_id"] == "MIR-001"
        assert configs[0]["cpu_model"] == "Intel Xeon E3-1231v3"
        assert configs[0]["cpu_sockets"] == 1
        assert configs[0]["cpu_cores_per_socket"] == 4
        assert configs[0]["ram_gb"] == 16
        assert configs[0]["miran_price"] == 8200.0
        assert configs[1]["cpu_sockets"] == 2
        assert len(configs[2]["disk_pools"]) == 2

    def test_config_without_price_kept(self, workbook):
        configs = convert(workbook, ALIASES)["configs"]
        assert configs[3]["cpu_model"] == "Intel Xeon E-2386G"
        assert configs[3]["miran_price"] is None

    def test_duplicate_config_skipped(self, workbook):
        configs = convert(workbook, ALIASES)["configs"]
        prices = [c["miran_price"] for c in configs
                  if c["cpu_model"] == "Intel Xeon Silver 4214R"]
        assert prices == [21000.0]  # вторая строка-дубликат (22222) не попала

    def test_source_rows_recorded(self, workbook):
        configs = convert(workbook, ALIASES)["configs"]
        assert [c["source_row"] for c in configs] == [3, 4, 6, 7]

    def test_missing_sheet_raises(self, workbook):
        with pytest.raises(ValueError, match="не найден"):
            convert(workbook, ALIASES, sheet="Нет_такого")


class TestApplyParserUpload:
    def test_upload_writes_both_configs(self, workbook, tmp_path, monkeypatch):
        import io
        import json

        import excel_to_configs as e2c
        import dedicated_app

        out = tmp_path / "miran_configs.json"
        classes_out = tmp_path / "disk_classes.json"
        monkeypatch.setattr(e2c, "DEFAULT_OUT", out)
        monkeypatch.setattr(e2c, "DEFAULT_CLASSES_OUT", classes_out)
        monkeypatch.setattr(dedicated_app, "DATA_DIR", str(tmp_path))
        # реальный словарь CPU: фикстура использует модели из cpu_specs.json
        n_cfg, n_groups, warns = dedicated_app.apply_parser_upload(
            io.BytesIO(workbook.read_bytes())
        )
        assert n_cfg == 4 and n_groups == 3
        assert json.loads(out.read_text(encoding="utf-8"))["configs"]
        assert json.loads(classes_out.read_text(encoding="utf-8"))["groups"]
        assert (tmp_path / "Parser.xlsx").exists()
        assert any("дубликат" in w for w in warns)

    def test_upload_without_mapping_sheet_rejected(self, tmp_path, monkeypatch):
        import io

        import openpyxl

        import excel_to_configs as e2c
        import dedicated_app
        import pytest

        monkeypatch.setattr(e2c, "DEFAULT_OUT", tmp_path / "m.json")
        monkeypatch.setattr(e2c, "DEFAULT_CLASSES_OUT", tmp_path / "d.json")
        monkeypatch.setattr(dedicated_app, "DATA_DIR", str(tmp_path))
        wb = openpyxl.Workbook()
        wb.active.title = "Данные"
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError):
            dedicated_app.apply_parser_upload(io.BytesIO(buf.getvalue()))
        assert not (tmp_path / "m.json").exists()


class TestConvertDiskClasses:
    def test_groups_by_section_and_units(self, workbook):
        groups = convert_disk_classes(workbook)["groups"]
        assert {"disk_type": "SSD", "sizes_gb": [960, 1000]} in groups
        assert {"disk_type": "NVMe", "sizes_gb": [375]} in groups
        assert {"disk_type": "HDD", "sizes_gb": [1000]} in groups
        assert len(groups) == 3

    def test_missing_sheet_raises(self, workbook):
        with pytest.raises(ValueError, match="не найден"):
            convert_disk_classes(workbook, sheet="Нет_такого")
