import shutil
import sys
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_export import (
    export_to_excel,
    remove_legacy_sheet,
    PROVIDER_COL,
    COLOR_COLS,
)

XLSX_SOURCE = Path(__file__).parent.parent / "data" / "Сравнение стоимости DS.xlsx"

TODAY = date(2026, 6, 15)


def _make_df():
    """Synthetic scrape result with a clear per-row min and max."""
    rows = [
        {
            "provider": "miran", "cpu_model": "Xeon E3-1231v3",
            "cpu_model_norm": "xeon e3-1231v3", "cpu_generation": "Haswell",
            "ram_gb": 16, "disk_count": 2, "disk_size_gb": 480, "disk_type": "SSD",
            "price_rub": 5000.0, "quantity_available": None, "scraped_at": "2026-06-15",
        },
        {
            "provider": "selectel", "cpu_model": "Xeon E3-1231v3",
            "cpu_model_norm": "xeon e3-1231v3", "cpu_generation": "Haswell",
            "ram_gb": 16, "disk_count": 2, "disk_size_gb": 480, "disk_type": "SSD",
            "price_rub": 9000.0, "quantity_available": None, "scraped_at": "2026-06-15",
        },
        {
            "provider": "1dedic", "cpu_model": "Xeon E3-1231v3",
            "cpu_model_norm": "xeon e3-1231v3", "cpu_generation": "Haswell",
            "ram_gb": 16, "disk_count": 2, "disk_size_gb": 480, "disk_type": "SSD",
            "price_rub": 7000.0, "quantity_available": None, "scraped_at": "2026-06-15",
        },
    ]
    return pd.DataFrame(rows)


@pytest.fixture
def xlsx_copy(tmp_path):
    """Provide a writable copy of the real workbook in a temp dir."""
    if not XLSX_SOURCE.exists():
        pytest.skip(f"Excel source not found: {XLSX_SOURCE}")
    dest = tmp_path / "test_copy.xlsx"
    shutil.copy(XLSX_SOURCE, dest)
    return str(dest)


class TestExportToExcel:
    def test_creates_new_sheet(self, xlsx_copy):
        sheet = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        wb = load_workbook(xlsx_copy)
        assert sheet in wb.sheetnames

    def test_sheet_name_format(self, xlsx_copy):
        sheet = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        assert re.match(r"^[А-Яа-яЁё]+_\d{2}_auto", sheet), f"unexpected name: {sheet}"

    def test_collision_gets_suffix(self, xlsx_copy):
        s1 = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        s2 = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        assert s1 != s2
        assert s2.endswith("_2")

    def test_existing_sheets_preserved(self, xlsx_copy):
        wb_before = load_workbook(xlsx_copy)
        original_sheets = set(wb_before.sheetnames)
        export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        wb_after = load_workbook(xlsx_copy)
        assert original_sheets.issubset(set(wb_after.sheetnames))

    def test_header_row_is_bold(self, xlsx_copy):
        sheet = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        wb = load_workbook(xlsx_copy)
        ws = wb[sheet]
        assert ws["A1"].font.bold is True
        assert ws["G1"].font.bold is True

    def test_miran_price_in_col_e(self, xlsx_copy):
        sheet = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        wb = load_workbook(xlsx_copy)
        ws = wb[sheet]
        miran_col = PROVIDER_COL["miran"]
        prices = [ws.cell(row=r, column=miran_col).value for r in range(2, ws.max_row + 1)]
        assert any(p == 5000.0 for p in prices)

    def test_calculator_col_f_empty(self, xlsx_copy):
        sheet = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        wb = load_workbook(xlsx_copy)
        ws = wb[sheet]
        for row in range(2, ws.max_row + 1):
            assert ws.cell(row=row, column=6).value in (None, "")

    def test_min_cell_gets_green_fill(self, xlsx_copy):
        sheet = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        wb = load_workbook(xlsx_copy)
        ws = wb[sheet]
        # miran=5000 is the row min → should be green
        miran_col = PROVIDER_COL["miran"]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=miran_col).value == 5000.0:
                fill = ws.cell(row=row, column=miran_col).fill
                assert fill.fgColor.rgb == "FFD9EAD3", f"expected green fill, got {fill.fgColor.rgb}"
                break

    def test_max_cell_gets_red_fill(self, xlsx_copy):
        sheet = export_to_excel(_make_df(), xlsx_path=xlsx_copy, today=TODAY)
        wb = load_workbook(xlsx_copy)
        ws = wb[sheet]
        # selectel=9000 is the row max → should be red
        selectel_col = PROVIDER_COL["selectel"]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=selectel_col).value == 9000.0:
                fill = ws.cell(row=row, column=selectel_col).fill
                assert fill.fgColor.rgb == "FFF4CCCC", f"expected red fill, got {fill.fgColor.rgb}"
                break


class TestRemoveLegacySheet:
    def test_removes_existing_sheet(self, xlsx_copy):
        wb = load_workbook(xlsx_copy)
        if "Лист 65" not in wb.sheetnames:
            pytest.skip("Лист 65 already removed from source")
        result = remove_legacy_sheet(xlsx_copy, name="Лист 65")
        assert result is True
        wb2 = load_workbook(xlsx_copy)
        assert "Лист 65" not in wb2.sheetnames

    def test_other_sheets_preserved_after_removal(self, xlsx_copy):
        wb = load_workbook(xlsx_copy)
        others = [s for s in wb.sheetnames if s != "Лист 65"]
        remove_legacy_sheet(xlsx_copy, name="Лист 65")
        wb2 = load_workbook(xlsx_copy)
        for s in others:
            assert s in wb2.sheetnames

    def test_idempotent_returns_false(self, xlsx_copy):
        remove_legacy_sheet(xlsx_copy, name="Лист 65")
        result2 = remove_legacy_sheet(xlsx_copy, name="Лист 65")
        assert result2 is False

    def test_nonexistent_sheet_returns_false(self, xlsx_copy):
        assert remove_legacy_sheet(xlsx_copy, name="ПривидениеНесуществующее") is False


import re  # noqa: E402 (import after test classes for clarity)
